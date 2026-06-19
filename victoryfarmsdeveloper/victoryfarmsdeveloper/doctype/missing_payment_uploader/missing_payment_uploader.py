# Copyright (c) 2026, Victory Farms and Contributors
# License: MIT. See LICENSE

import frappe
from frappe.model.document import Document
import openpyxl
import io
import os
from datetime import datetime


class MissingPaymentUploader(Document):
    def validate(self):
        if self.payment_file and not self.result_log:
            frappe.msgprint(
                "Please click 'Process Payments' after saving to upload the data.",
                alert=True,
            )

    @frappe.whitelist()
    def process_payments(self):
        """Read the uploaded Excel file and create Mpesa Payment Register records."""
        if not self.payment_file:
            frappe.throw("Please upload an Excel file first.")

        self.status = "Processing"
        self.save(ignore_permissions=True)
        frappe.db.commit()

        # Column mapping: Excel header name -> Mpesa Payment Register fieldname
        COLUMN_MAP = {
            "First Name": "first_name",
            "Full Name": "full_name",
            "Business Short Code": "business_short_code",
            "MSISDN": "msisdn",
            "Trans ID": "transaction_id",
            "Trans Time": "trans_time",
            "Trans Amount": "trans_amount",
            "Transaction Type": "transaction_type",
            "Posting Date": "posting_date",
            "Posting Time": "posting_time",
            "Company": "company",
            "Default Currency": "default_currency",
            "Currency": "currency",
        }

        # Fieldnames that link to other DocTypes (must be resolved)
        LINK_FIELDS = {
            "company": "Company",
            "customer": "Customer",
            "mode_of_payment": "Mode of Payment",
        }

        try:
            file_url = self.payment_file
            file_path = frappe.get_site_path(
                "private" + file_url if file_url.startswith("/files/") else file_url
            )

            # Handle files uploaded via Attach field
            if file_url.startswith("/files/"):
                file_path = frappe.get_site_path("private" + file_url)
            elif file_url.startswith("/private/files/"):
                file_path = frappe.get_site_path(file_url)
            else:
                # Try the files folder
                file_path = frappe.get_site_path("private/files", os.path.basename(file_url))

            # If path doesn't exist, try getting it via frappe.get_doc
            if not os.path.exists(file_path):
                file_doc = frappe.get_doc("File", {"file_url": file_url})
                file_path = frappe.get_site_path(
                    "private" + file_doc.file_url
                    if file_doc.file_url.startswith("/files/")
                    else file_doc.file_url
                )

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Read headers from first row
            headers = [cell.value for cell in ws[1] if cell.value is not None]
            if not headers:
                frappe.throw("The Excel file appears to be empty or has no headers.")

            # Map Excel columns to fieldnames
            col_to_field = {}
            unmapped = []
            for idx, header in enumerate(headers):
                header_clean = str(header).strip()
                if header_clean in COLUMN_MAP:
                    col_to_field[idx] = COLUMN_MAP[header_clean]
                else:
                    unmapped.append(header_clean)

            if not col_to_field:
                frappe.throw(
                    f"No matching columns found. Expected headers: {', '.join(COLUMN_MAP.keys())}"
                )

            success_count = 0
            failure_count = 0
            total_rows = 0
            errors = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(v is None for v in row):
                    continue  # Skip empty rows

                total_rows += 1
                row_data = {}

                for col_idx, fieldname in col_to_field.items():
                    if col_idx < len(row):
                        value = row[col_idx]
                        if value is not None:
                            row_data[fieldname] = value

                if not row_data:
                    continue

                # Convert posting_date to proper format
                if "posting_date" in row_data:
                    val = row_data["posting_date"]
                    if isinstance(val, datetime):
                        row_data["posting_date"] = val.strftime("%Y-%m-%d")
                    elif isinstance(val, str):
                        # Try parsing common date formats
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                            try:
                                row_data["posting_date"] = datetime.strptime(
                                    val, fmt
                                ).strftime("%Y-%m-%d")
                                break
                            except ValueError:
                                continue

                # Ensure posting_date is always set
                if "posting_date" not in row_data:
                    row_data["posting_date"] = frappe.utils.today()

                # Resolve Link fields
                for fieldname, doctype in LINK_FIELDS.items():
                    if fieldname in row_data:
                        val = str(row_data[fieldname]).strip()
                        if val and not frappe.db.exists(doctype, val):
                            # Try finding by name
                            result = frappe.db.get_value(doctype, val, "name")
                            if not result:
                                errors.append(
                                    f"Row {row_num}: {doctype} '{val}' not found. Skipping {fieldname}."
                                )
                                row_data.pop(fieldname, None)
                            else:
                                row_data[fieldname] = result
                        elif not val:
                            row_data.pop(fieldname, None)

                # Create the Mpesa Payment Register document
                try:
                    mpesa_doc = frappe.get_doc(
                        {"doctype": "Mpesa Payment Register", **row_data}
                    )
                    mpesa_doc.insert(ignore_permissions=True)
                    success_count += 1
                except Exception as e:
                    failure_count += 1
                    errors.append(f"Row {row_num}: {str(e)}")

            wb.close()

            # Build result log
            log_lines = []
            log_lines.append(f"=== Import Summary ===")
            log_lines.append(f"Total data rows: {total_rows}")
            log_lines.append(f"Successfully imported: {success_count}")
            log_lines.append(f"Failed: {failure_count}")
            if unmapped:
                log_lines.append(f"\nUnmapped columns (ignored): {', '.join(unmapped)}")
            if errors:
                log_lines.append(f"\n--- Error Details ---")
                log_lines.extend(errors)

            self.total_rows = total_rows
            self.success_count = success_count
            self.failure_count = failure_count
            self.result_log = "\n".join(log_lines)
            self.status = "Completed" if failure_count == 0 else "Completed"
            self.save(ignore_permissions=True)
            frappe.db.commit()

            return {
                "success": True,
                "total": total_rows,
                "imported": success_count,
                "failed": failure_count,
                "message": f"Import complete. {success_count} of {total_rows} records imported successfully.",
            }

        except Exception as e:
            self.status = "Failed"
            self.result_log = f"Import failed with error:\n{str(e)}"
            self.save(ignore_permissions=True)
            frappe.db.commit()
            frappe.log_error(
                title="Missing Payment Uploader Error", message=str(e)
            )
            return {"success": False, "message": str(e)}