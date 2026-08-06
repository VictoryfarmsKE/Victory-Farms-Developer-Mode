import frappe
from frappe import _


PO_REF_SEPARATOR = " | "


def get_po_refs(doc):
    """Return a list of unique Purchase Order reference names."""
    if not doc.references:
        return []

    po_refs = [
        r.reference_name
        for r in doc.references
        if r.reference_doctype == "Purchase Order" and r.reference_name
    ]
    return list(dict.fromkeys(po_refs))  # preserve order, remove duplicates


def set_beneficiary_purpose_of_payment(doc, method=None):
    """Auto-fill PurposeOfPayment on all beneficiary rows from PO references.

    The PO reference(s) are stored in pipe-separated form to match the M-Pesa
    export format. When multiple POs exist they are joined with ' | '.
    """
    po_refs = get_po_refs(doc)
    new_value = PO_REF_SEPARATOR.join(po_refs) if po_refs else ""

    # Auto-fill Reference Number with the first PO reference
    if po_refs and doc.reference_no != po_refs[0]:
        doc.reference_no = po_refs[0]

    if not doc.get("custom_beneficiaries"):
        return

    for row in doc.custom_beneficiaries:
        if row.purpose_of_payment != new_value:
            row.purpose_of_payment = new_value


def validate_beneficiary_fields(doc, method=None):
    """Server-side validation for beneficiary rows.

    At least one beneficiary row is required, and each row must contain all
    M-Pesa upload fields.
    """
    if not doc.get("custom_beneficiaries"):
        frappe.throw(_("At least one Beneficiary row is required on Payment Entry."))

    required_fields = {
        "MobileNumber": "mobile_number",
        "DocumentType": "document_type",
        "DocumentNumber": "document_number",
        "Amount": "amount",
        "PurposeOfPayment": "purpose_of_payment",
    }

    for idx, row in enumerate(doc.custom_beneficiaries, start=1):
        missing = [
            label for label, fieldname in required_fields.items() if not row.get(fieldname)
        ]
        if missing:
            frappe.throw(
                _(
                    "Row {0} in Beneficiaries is missing the following mandatory fields: {1}"
                ).format(idx, ", ".join(missing))
            )


@frappe.whitelist()
def upload_beneficiaries(file_url):
    """Parse an uploaded Excel/CSV file and return beneficiary rows.

    Expected columns (case-insensitive): MobileNumber, DocumentType,
    DocumentNumber, Amount, PurposeOfPayment, Name.
    Name is optional; PurposeOfPayment is auto-filled from the Payment
    Entry's Purchase Order references if not provided in the upload.
    """
    import csv
    import io
    import os

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()
    filename = file_doc.file_name or ""
    ext = os.path.splitext(filename)[1].lower()

    rows = []

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            frappe.throw(_("openpyxl is required to read Excel files."))

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_map = _build_header_map(headers)

        for excel_row in sheet.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in excel_row):
                continue
            mapped = {key: _get_cell_value(excel_row, idx) for key, idx in header_map.items()}
            rows.append(mapped)

    elif ext == ".csv":
        with open(file_path, mode="r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            headers = [h.strip() for h in next(reader)]
            header_map = _build_header_map(headers)
            for csv_row in reader:
                if not csv_row or all(v.strip() == "" for v in csv_row):
                    continue
                mapped = {key: _get_cell_value(csv_row, idx) for key, idx in header_map.items()}
                rows.append(mapped)
    else:
        frappe.throw(_("Unsupported file type. Please upload a .xlsx, .xls or .csv file."))

    # Normalize and validate rows
    normalized_rows = []
    for idx, row in enumerate(rows, start=2):
        mobile_number = _clean_value(row.get("mobile_number"))
        document_type = _clean_value(row.get("document_type"))
        document_number = _clean_value(row.get("document_number"))
        amount = _clean_value(row.get("amount"))
        beneficiary_name = _clean_value(row.get("beneficiary_name"))

        if not any([mobile_number, document_number, amount]):
            continue

        normalized_rows.append(
            {
                "mobile_number": mobile_number,
                "document_type": document_type or "National Id",
                "document_number": document_number,
                "amount": flt(amount),
                "beneficiary_name": beneficiary_name,
                "purpose_of_payment": _clean_value(row.get("purpose_of_payment")),
            }
        )

    return normalized_rows


def _build_header_map(headers):
    """Map expected column names to zero-based indices."""
    expected = {
        "mobile_number": ["mobilenumber", "mobile number", "mobile_number"],
        "document_type": ["documenttype", "document type", "document_type"],
        "document_number": ["documentnumber", "document number", "document_number"],
        "amount": ["amount"],
        "purpose_of_payment": ["purposeofpayment", "purpose of payment", "purpose_of_payment"],
        "beneficiary_name": ["name", "beneficiary_name", "beneficiary name"],
    }
    header_map = {}
    for idx, header in enumerate(headers):
        normalized = header.lower().replace(" ", "_").replace("-", "_")
        for key, aliases in expected.items():
            if normalized in aliases:
                header_map[key] = idx
                break
    return header_map


def _get_cell_value(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def _clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def flt(value, precision=None):
    from frappe.utils import flt as _flt

    return _flt(value, precision)


