"""
Generate an Excel file of Victory Farms warehouses with classification.

Columns:
  - Warehouse Name
  - Status (ENABLED / DISABLED)
  - Fixed Valuation Zone (1 = FLC/branch, 0 = internal/operational)

Run locally:  python scripts/generate_warehouse_excel.py
Requires:     pip install openpyxl
Output:       scripts/warehouse_classification.xlsx
"""

import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl is required. Install it with: pip install openpyxl")


# (warehouse_name, status) — exactly as exported from ERPNext
WAREHOUSES = [
    ("Feed Internal Transit - VFL", "ENABLED"),
    ("Feed Warehouse Roo - VFL", "ENABLED"),
    ("Feeds Store - VFL", "DISABLED"),
    ("Fish Q - VFL", "ENABLED"),
    ("Fish Q VF Store - VFL", "DISABLED"),
    ("FishQ Bolt - VFL", "DISABLED"),
    ("FishQ Procurement - VFL", "DISABLED"),
    ("FLC - VFL", "ENABLED"),
    ("FLC Branch Returns - VFL", "DISABLED"),
    ("FLC Branch Spoilage - VFL", "ENABLED"),
    ("FLC Deformed - VFL", "ENABLED"),
    ("FLC Donations and Samples - VFL", "ENABLED"),
    ("FLC Spoilage - VFL", "ENABLED"),
    ("FLC Variance - VFL", "ENABLED"),
    ("Free Area - VFL", "DISABLED"),
    ("Free Area - VFL Spoilage", "ENABLED"),
    ("Fuel-Inventory - VFL", "ENABLED"),
    ("Garden Estate - VFL", "ENABLED"),
    ("Garden Estate - VFL Spoilage", "ENABLED"),
    ("Githurai - VFL", "ENABLED"),
    ("Githurai - VFL Spoilage", "ENABLED"),
    ("Gobbi Concession - VFL", "ENABLED"),
    ("Grader Feed Warehouse - VFL", "ENABLED"),
    ("Hatchery - VFL", "DISABLED"),
    ("Hatchery Converted Feed - VFL", "ENABLED"),
    ("Hatchery Feed - VFL", "DISABLED"),
    ("Hatchery Store - VFL", "ENABLED"),
    ("Homa Bay Region - VFL", "ENABLED"),
    ("Homabay - VFL", "ENABLED"),
    ("Homabay - VFL Spoilage", "ENABLED"),
    ("Huruma - VFL", "ENABLED"),
    ("Huruma - VFL Spoilage", "ENABLED"),
    ("Imara - VFL", "ENABLED"),
    ("Imara - VFL Spoilage", "ENABLED"),
    ("Import Warehouse - VFL", "ENABLED"),
    ("In Transit - VFL", "ENABLED"),
    ("In Transit Variance - VFL", "ENABLED"),
    ("Isebania - VFL", "ENABLED"),
    ("Isebania - VFL Spoilage", "ENABLED"),
    ("IT Store - VFL", "ENABLED"),
    ("Jogoo Road - VFL", "ENABLED"),
    ("Jogoo Road - VFL Spoilage", "ENABLED"),
    ("Juja - VFL", "ENABLED"),
    ("Juja - VFL Spoilage", "ENABLED"),
    ("Kahawa Wendani - VFL", "ENABLED"),
    ("Kahawa Wendani - VFL Spoilage", "ENABLED"),
    ("Kahawa West - VFL", "ENABLED"),
    ("Kahawa West - VFL Spoilage", "ENABLED"),
    ("Kakamega - VFL", "ENABLED"),
    ("Kakamega - VFL Spoilage", "ENABLED"),
    ("Kakamega Region - VFL", "ENABLED"),
    ("Kalahari Concession - VFL", "ENABLED"),
    ("Kanga 5MT-KDX 985Z - VFL", "ENABLED"),
    ("Kangemi - VFL", "ENABLED"),
    ("Kangemi - VFL Spoilage", "ENABLED"),
    ("Kangundo Road - VFL", "ENABLED"),
    ("Kangundo Road - VFL Spoilage", "ENABLED"),
    ("Kariobangi North - VFL", "DISABLED"),
    ("Kariobangi North - VFL Spoilage", "ENABLED"),
    ("Kariobangi South - VFL", "ENABLED"),
    ("Kariobangi South - VFL Spoilage", "ENABLED"),
    ("Kasarani - VFL", "ENABLED"),
    ("Kasarani - VFL Spoilage", "ENABLED"),
    ("Katito - VFL", "ENABLED"),
    ("Katito - VFL Spoilage", "ENABLED"),
    ("Kawangware - VFL", "ENABLED"),
    ("Kawangware - VFL Spoilage", "ENABLED"),
    ("Kayole - VFL", "ENABLED"),
    ("Kayole - VFL Spoilage", "ENABLED"),
    ("KDH 223R-HIACE - VFL", "ENABLED"),
    ("KDK040Q - VFL", "ENABLED"),
    ("KDK404Q - VFL", "ENABLED"),
    ("Kehancha - VFL", "ENABLED"),
    ("Kehancha - VFL Spoilage", "ENABLED"),
    ("Kericho Region - VFL", "ENABLED"),
    ("Keroka - VFL", "ENABLED"),
    ("Keroka - VFL Spoilage", "ENABLED"),
    ("Keyo Fresh Logistics 5MT KDC 891C - VFL", "ENABLED"),
    ("Keyo Fresh Logistics 5MT KDU 402W - VFL", "ENABLED"),
    ("Kibera - VFL", "DISABLED"),
    ("Kibera - VFL Spoilage", "DISABLED"),
    ("Kiboko 3.5Mt Kcs 559A - VFL", "ENABLED"),
    ("Kifaru 13.5Mt Kcs 560A - VFL", "ENABLED"),
    ("Kimilili - VFL", "ENABLED"),
    ("Kimilili Spoilage - VFL", "ENABLED"),
    ("Kisii - Keroka - VFL", "ENABLED"),
    ("Kisii - Keroka - VFL Spoilage", "ENABLED"),
    ("Kisii - VFL", "ENABLED"),
    ("Kisii - VFL Spoilage", "ENABLED"),
    ("Kisii Region - VFL", "ENABLED"),
    ("Kisumu Hatchery - VFL", "DISABLED"),
    ("Kisumu Region - VFL", "ENABLED"),
    ("Kitengela - VFL", "ENABLED"),
    ("Kitengela - VFL Spoilage", "ENABLED"),
    ("KLC - VFL", "ENABLED"),
    ("KLC Branch Returns - VFL", "DISABLED"),
    ("KLC Branch Spoilage - VFL", "ENABLED"),
    ("KLC Consumables Store - VFL", "ENABLED"),
    ("KLC Deformed - VFL", "ENABLED"),
    ("KLC Donations and Samples - VFL", "ENABLED"),
    ("KLC Procurement - VFL", "DISABLED"),
    ("KLC Sales - VFL", "ENABLED"),
    ("KLC Spoilage - VFL", "ENABLED"),
    ("KLC Variance - VFL", "ENABLED"),
    ("KOBE KDS 412X - VFL", "ENABLED"),
    ("Kondele - VFL", "ENABLED"),
    ("Kondele - VFL Spoilage", "ENABLED"),
    ("Kosodo Concession  - VFL", "ENABLED"),
    ("KSM - VFL", "ENABLED"),
    ("Lake - VFL", "ENABLED"),
    ("Lake Test - VFL", "DISABLED"),
    ("Lakeshore - VFL", "ENABLED"),
    ("Lakeshore Test - VFL", "DISABLED"),
    ("Logistics Warehouse - VFL", "DISABLED"),
    ("Luanda - VFL", "ENABLED"),
    ("Luanda - VFL Spoilage", "ENABLED"),
    ("Lucky Summer - VFL", "ENABLED"),
    ("Lucky Summer - VFL Spoilage", "ENABLED"),
    ("M/T1 Bamburi Branch - VFL", "ENABLED"),
    ("M/T1 Bamburi Branch - VFL Spoilage", "ENABLED"),
    ("M/T1 Kisauni - VFL", "ENABLED"),
    ("M/T1 Kisauni - VFL Spoilage", "ENABLED"),
    ("M/T1 Kongowea - VFL", "ENABLED"),
    ("M/T1 Kongowea - VFL Spoilage", "ENABLED"),
    ("M/T1 Majengo-Sakina Branch - VFL", "ENABLED"),
    ("M/T1 Majengo-Sakina Branch - VFL Spoilage", "ENABLED"),
    ("M/T1 Mombasa-Bombolulu  - VFL", "ENABLED"),
    ("M/T1 Mombasa-Bombolulu  - VFL Spoilage", "ENABLED"),
    ("M/T1 Mshomoroni Branch - VFL", "ENABLED"),
    ("M/T1 Mshomoroni Branch - VFL Spoilage", "ENABLED"),
    ("M/T1 Mtwapa Branch - VFL", "ENABLED"),
    ("M/T1 Mtwapa Branch - VFL Spoilage", "ENABLED"),
    ("M/T1 Tononoka - VFL", "ENABLED"),
    ("M/T1 Tononoka - VFL Spoilage", "ENABLED"),
    ("Machakos - VFL", "ENABLED"),
    ("Machakos - VFL Spoilage", "ENABLED"),
    ("Majengo - VFL", "ENABLED"),
    ("Majengo - VFL Spoilage", "ENABLED"),
    ("Makongeni - VFL", "DISABLED"),
    ("Makongeni - VFL Spoilage", "DISABLED"),
    ("Mamba 2.8Mt Kcq 982Z - VFL", "ENABLED"),
    ("Manyatta - VFL", "ENABLED"),
    ("Manyatta - VFL Spoilage", "ENABLED"),
    ("Masara - VFL", "ENABLED"),
    ("Masara - VFL Spoilage", "ENABLED"),
    ("Materials in transit-Feed - VFL", "ENABLED"),
    ("Materials in transit-Other - VFL", "ENABLED"),
    ("Mathare - VFL", "ENABLED"),
    ("Mathare - VFL Spoilage", "ENABLED"),
    ("Mawenzi KDL 168G - VFL", "ENABLED"),
    ("Mbale - VFL", "ENABLED"),
    ("Mbale - VFL Spoilage", "ENABLED"),
    ("Mbita - VFL", "ENABLED"),
    ("Mbita - VFL Spoilage", "ENABLED"),
    ("MBUNI 28MT KDX 689L - VFL", "ENABLED"),
    ("Meru - VFL", "ENABLED"),
    ("Meru Consumables - VFL", "ENABLED"),
    ("Meru LC Sales - VFL", "ENABLED"),
    ("Migori - VFL", "ENABLED"),
    ("Migori - VFL Spoilage", "ENABLED"),
    ("Migori Region - VFL", "ENABLED"),
    ("Mini - Store 14 - VFL", "DISABLED"),
    ("Mini - Store Hatchery 2 - VFL", "ENABLED"),
    ("Mini Store 24 - VFL", "ENABLED"),
    ("Mini- Store 20 - VFL", "ENABLED"),
    ("Mini- Store 21 - VFL", "ENABLED"),
    ("Mini- Store 25 - VFL", "ENABLED"),
    ("Mini- Store 26 - VFL", "ENABLED"),
    ("Mini-Store 10 - VFL", "ENABLED"),
    ("Mini-Store 12 - VFL", "ENABLED"),
    ("Mini-Store 14 - VFL", "ENABLED"),
    ("Mini-Store 15 - VFL", "ENABLED"),
    ("Mini-Store 16 - VFL", "ENABLED"),
    ("Mini-Store 18 - VFL", "ENABLED"),
    ("Mini-Store 19 - VFL", "ENABLED"),
    ("Mini-Store 22 - VFL", "ENABLED"),
    ("Mini-Store 23 - VFL", "ENABLED"),
    ("Mini-Store 3 - VFL", "ENABLED"),
    ("Mini-Store 4 - VFL", "ENABLED"),
    ("Mini-Store 5 - VFL", "ENABLED"),
    ("Mini-Store 7 - VFL", "ENABLED"),
    ("Mini-Store 9 - VFL", "ENABLED"),
    ("Mini-Store Foodstuffs - VFL", "ENABLED"),
    ("Mini-Store Hatchery 1 - VFL", "ENABLED"),
    ("Mini-Store Hazardous - VFL", "ENABLED"),
    ("Mini-Store Lakeshore - VFL", "ENABLED"),
    ("Mini-Store Net Area - VFL", "ENABLED"),
    ("MK/T2 - Chuka Branch - VFL", "ENABLED"),
    ("MK/T2 - Chuka Spoilage", "ENABLED"),
    ("MK/T2 - Kaaga Branch - VFL", "ENABLED"),
    ("MK/T2 - Kaaga Spoilage", "ENABLED"),
    ("MK/T2 - Meru CBD Branch - VFL", "ENABLED"),
    ("MK/T2 - Meru CBD Spoilage", "ENABLED"),
    ("MK/T2 - Meru-LC  Branch Spoilage", "ENABLED"),
    ("MK/T2 - Meru-LC - VFL", "ENABLED"),
    ("MK/T2 - Meru-LC Deformed - VFL", "ENABLED"),
    ("MK/T2 - Meru-LC Donations and Samples", "ENABLED"),
    ("MK/T2 - Meru-LC Spoilage", "ENABLED"),
    ("MK/T2 - Meru-LC Variance", "ENABLED"),
    ("MK/T2 - Nkubu Town Branch - VFL", "ENABLED"),
    ("MK/T2 - Nkubu Town Spoilage", "ENABLED"),
    ("MK/T3 - Chogoria Branch - VFL", "ENABLED"),
    ("MK/T3 - Chogoria Spoilage", "ENABLED"),
    ("MK/T3 - Maua Branch - VFL", "ENABLED"),
    ("MK/T3 - Maua Spoilage", "ENABLED"),
    ("MK/T3 - Muthara Branch - VFL", "ENABLED"),
    ("MK/T3 - Muthara Spoilage", "ENABLED"),
    ("Mlango Kubwa - VFL", "DISABLED"),
    ("Mlango Kubwa - VFL Spoilage", "DISABLED"),
    ("Mlango Nandi - VFL", "ENABLED"),
    ("Mlango Nandi Spoilage - VFL", "ENABLED"),
    ("MLC - VFL", "ENABLED"),
    ("MLC Branch Returns - VFL", "DISABLED"),
    ("MLC Branch Spoilage - VFL", "ENABLED"),
    ("MLC Consumables Store - VFL", "ENABLED"),
    ("MLC Donations and Samples - VFL", "ENABLED"),
    ("MLC Procurement - VFL", "DISABLED"),
    ("MLC Returns - VFL", "DISABLED"),
    ("MLC Sales - VFL", "ENABLED"),
    ("MLC Spoilage - VFL", "ENABLED"),
    ("MLC Variance - VFL", "ENABLED"),
    ("Mlolongo - VFL", "ENABLED"),
    ("Mlolongo - VFL Spoilage", "ENABLED"),
    ("Mombasa - VFL", "ENABLED"),
    ("Mount Kenya - VFL", "ENABLED"),
    ("Muhoroni - VFL", "ENABLED"),
    ("Muhoroni - VFL Spoilage", "ENABLED"),
    ("Mumias - VFL", "ENABLED"),
    ("Mumias - VFL Spoilage", "ENABLED"),
    ("Mwewe 3.5Mt Kct 213T - VFL", "ENABLED"),
    ("Mwiki - VFL", "ENABLED"),
    ("Mwiki - VFL Spoilage", "ENABLED"),
    ("Nairobi West - VFL", "ENABLED"),
    ("Nairobi West - VFL Spoilage", "ENABLED"),
    ("Naivas Capital - VFL", "ENABLED"),
    ("Naivas Capital - VFL Spoilage", "ENABLED"),
    ("Nakuru", "ENABLED"),
    ("Nandi Hills - VFL", "ENABLED"),
    ("Nandi Hills Spoilage - VFL", "ENABLED"),
    ("NBO - VFL", "ENABLED"),
    ("Ndhiwa - VFL", "ENABLED"),
    ("Ndhiwa - VFL Spoilage", "ENABLED"),
    ("Ndovu 28Mt Kcy 048N - VFL", "ENABLED"),
    ("Ngara - VFL", "ENABLED"),
    ("Ngara - VFL Spoilage", "ENABLED"),
    ("Ngong Town - VFL", "ENABLED"),
    ("Ngong Town - VFL Spoilage", "ENABLED"),
    ("Njiwa 5MT KDX 984X  - VFL", "ENABLED"),
    ("North America - VFL", "ENABLED"),
    ("North Region - VFL", "DISABLED"),
    ("NQR81 BLUEPOWER TRUCK 1 - VFL", "ENABLED"),
    ("NQR81 BLUEPOWER TRUCK 2 - VFL", "ENABLED"),
    ("Nyakoe - VFL", "ENABLED"),
    ("Nyakoe - VFL Spoilage", "ENABLED"),
    ("Nyalenda - VFL", "ENABLED"),
    ("Nyalenda - VFL Spoilage", "ENABLED"),
    ("Nyamasaria - VFL", "ENABLED"),
    ("Nyamasaria - VFL Spoilage", "ENABLED"),
    ("Nyamira - VFL", "ENABLED"),
    ("Nyamira - VFL Spoilage", "ENABLED"),
    ("Nyati 28Mt KDP 369K-VLC - VFL", "DISABLED"),
    ("NYATI 28MT KDP 396K-VLC - VFL", "ENABLED"),
    ("Nyiri Concession  - VFL", "ENABLED"),
    ("Oceania - VFL", "ENABLED"),
    ("Operations - VFL", "DISABLED"),
    ("Oremo Early stage Concession - VFL", "ENABLED"),
    ("Otonglo - Spoilage - VFL", "ENABLED"),
    ("Otonglo - VFL", "ENABLED"),
    ("Oyugis - VFL", "ENABLED"),
    ("Oyugis - VFL Spoilage", "ENABLED"),
    ("Paa 5Mt KDU 086V - VFL", "ENABLED"),
    ("Pipelihne Juakali Spoilage - VFL", "ENABLED"),
    ("Pipeline - VFL", "ENABLED"),
    ("Pipeline - VFL Spoilage", "ENABLED"),
    ("Pipeline- Juakali - VFL", "ENABLED"),
    ("Pond - VFL", "ENABLED"),
    ("Ponds Feeds Warehouse - VFL", "ENABLED"),
    ("Processing - VFL", "ENABLED"),
    ("Procurement - VFL", "DISABLED"),
    ("Production - VFL", "DISABLED"),
    ("Pwani  - VFL", "ENABLED"),
    ("Pweza 5Mt KDU 084V - VFL", "ENABLED"),
    ("Rejected Warehouse - VFL", "ENABLED"),
    ("RIDER - Brian - VFL", "ENABLED"),
    ("Rift V", "ENABLED"),
    ("Rift V - VFL", "ENABLED"),
    ("Rift VAL", "ENABLED"),
    ("Rift Valley - VFL", "ENABLED"),
    ("Rodi - VFL", "ENABLED"),
    ("Rodi - VFL Spoilage", "ENABLED"),
    ("Rongai - VFL", "ENABLED"),
    ("Rongai - VFL Spoilage", "ENABLED"),
    ("Rongo - VFL", "ENABLED"),
    ("Rongo - VFL Spoilage", "ENABLED"),
    ("Roo - VFL", "ENABLED"),
    ("Roo - VFL Spoilage", "ENABLED"),
    ("Roo Farm On Shore - VFL", "DISABLED"),
    ("Roo Farm Store - VFL", "ENABLED"),
    ("Roo Trials Warehouse - VFL", "ENABLED"),
    ("Roo Truck - VFL", "ENABLED"),
    ("Ruai - VFL", "ENABLED"),
    ("Ruai - VFL Spoilage", "ENABLED"),
    ("Sales - VFL", "ENABLED"),
    ("Satellite - VFL", "ENABLED"),
    ("Satellite - VFL Spoilage", "ENABLED"),
    ("Shabab - VFL", "ENABLED"),
    ("Shabab - VFL Spoilage", "ENABLED"),
    ("Shopify Warehouse - VFL", "ENABLED"),
    ("Siaya - VFL", "DISABLED"),
    ("Siaya - VFL Spoilage", "ENABLED"),
    ("Siaya Region - VFL", "DISABLED"),
    ("Simba 13.5Mt Kct 466T - VFL", "ENABLED"),
    ("Sindo - VFL", "ENABLED"),
    ("Sindo - VFL Spoilage", "ENABLED"),
    ("Skifoc 5MT KDA 864L - VFL", "ENABLED"),
    ("SKIFOC 5MT KDK 404Q - VFL", "ENABLED"),
    ("Skifoc 5T KDH 298R - VFL", "ENABLED"),
    ("Sondu - VFL", "ENABLED"),
    ("Sondu - VFL Spoilage", "ENABLED"),
    ("South America - VFL", "ENABLED"),
    ("South Region - VFL", "DISABLED"),
    ("Suna East - VFL", "ENABLED"),
    ("Suna East - VFL Spoilage", "ENABLED"),
    ("Suneka - VFL", "ENABLED"),
    ("Suneka - VFL Spoilage", "ENABLED"),
    ("Support - VFL", "DISABLED"),
    ("Swara 4.5Mt Kdc 094U - VFL", "ENABLED"),
    ("TAI 28MT KDS 422U  - VFL", "ENABLED"),
    ("Tatacoa Concession  - VFL", "ENABLED"),
    ("Test 234s - VFL", "DISABLED"),
    ("Test Warehouse - VFL", "DISABLED"),
    ("Test Warehouse 540001 - Biological assets - VFL", "DISABLED"),
    ("Test Warehouse 540003 - Gutted Tilapia Fish - VFL", "DISABLED"),
    ("Test Warehouse 540004 - Feed-Inventory - VFL", "DISABLED"),
    ("Test Warehouse 540005 - Fuel-Inventory - VFL", "DISABLED"),
    ("Test Warehouse 540006 - Consumables - VFL", "DISABLED"),
    ("Test Warehouse 540008 - Materials in transit-Feed - VFL", "DISABLED"),
    ("Test Warehouse 540009 - Materials in transit-Other - VFL", "DISABLED"),
    ("Thika - VFL", "ENABLED"),
    ("Thika - VFL Spoilage", "ENABLED"),
    ("Toyota Hilux Z Edition (2022) (Black) - VFL", "ENABLED"),
    ("TPL-KAI-5MT KDB 358Y - VFL", "ENABLED"),
    ("TPL-KAI-5Mt KDR 358Y - VFL", "ENABLED"),
    ("TPL-SK-KDE 127R - VFL", "DISABLED"),
    ("Twiga 3.5Mt Kcs 561A - VFL", "ENABLED"),
    ("Ugunja - VFL", "ENABLED"),
    ("Ugunja - VFL Spoilage", "ENABLED"),
    ("Umoja - VFL", "ENABLED"),
    ("Umoja - VFL Spoilage", "ENABLED"),
    ("Utawala - VFL", "ENABLED"),
    ("Utawala - VFL Spoilage", "ENABLED"),
    ("VF Store - VFL", "DISABLED"),
    ("Victory Fresh - VFL", "ENABLED"),
    ("Victory Fresh - VFL Spoilage", "ENABLED"),
    ("VLC - Customer Direct - VFL", "ENABLED"),
    ("VLC - VFL", "ENABLED"),
    ("VLC Admin Store - VFL", "ENABLED"),
    ("VLC Branch Returns - VFL", "DISABLED"),
    ("VLC Branch Spoilage - VFL", "ENABLED"),
    ("VLC Consumables Store  - VFL", "ENABLED"),
    ("VLC Deformed - VFL", "ENABLED"),
    ("VLC Donations and Samples - VFL", "ENABLED"),
    ("VLC Feed Store - VFL", "ENABLED"),
    ("VLC Sales - VFL", "ENABLED"),
    ("VLC Sales - VFL Spoilage", "ENABLED"),
    ("VLC Spoilage - VFL", "ENABLED"),
    ("VLC Store Procurement - VFL", "DISABLED"),
    ("VLC Variance - VFL", "ENABLED"),
    ("W/T3 - Siaya - VFL", "ENABLED"),
    ("W/T4 Kendu Bay Branch - VFL", "ENABLED"),
    ("Webuye - VFL", "ENABLED"),
    ("Webuye Spoilage - VFL", "ENABLED"),
    ("West Region - VFL", "DISABLED"),
    ("Western  - VFL", "ENABLED"),
    ("Yaya - VFL", "ENABLED"),
    ("Yaya - VFL Spoilage", "ENABLED"),
    ("Zimmerman - VFL", "ENABLED"),
    ("Zimmerman - VFL Spoilage", "ENABLED"),
]


# ------------------------------------------------------------------
# Classification logic
#   1 = FLC + branch (fixed valuation zone)
#   0 = internal / operational / logistics / vehicles / test / group
# ------------------------------------------------------------------

# Warehouses that are 0 even though they look like LCs/branches (group nodes,
# channels, internal stores, procurement, etc.)
FORCE_ZERO = {
    "Processing - VFL",
    "Sales - VFL",
    "NBO - VFL",
    "KSM - VFL",
    "Lake - VFL",
    "Pond - VFL",
    "Lakeshore - VFL",
    "Roo Farm On Shore - VFL",
    "Mount Kenya - VFL",
    "Rift Valley - VFL",
    "Rift V - VFL",
    "Rift V",
    "Rift VAL",
    "Nakuru",
    "North America - VFL",
    "South America - VFL",
    "Oceania - VFL",
    "Shopify Warehouse - VFL",
    "Victory Fresh - VFL",
    "Victory Fresh - VFL Spoilage",
    "VLC - Customer Direct - VFL",
    "Rejected Warehouse - VFL",
}

# Keyword fragments (lowercase) that always mean 0 (internal/operational)
ZERO_KEYWORDS = [
    "procurement",
    "consumables store",
    "consumables - vfl",
    "admin store",
    "feed store",
    "feeds store",
    "vf store",
    "it store",
    "feed warehouse",
    "feeds warehouse",
    "feed internal transit",
    "hatchery",
    "broodstock",
    "farm store",
    "farm central store",
    "farm kitchen",
    "production",
    "operations",
    "support",
    "construction",
    "import warehouse",
    "logistics warehouse",
    "in transit",
    "materials in transit",
    "fuel-inventory",
    "mini-store",
    "mini - store",
    "mini store",
    "mini- store",
    "concession",
    "trials warehouse",
    "grader feed",
    "ponds feeds",
    "fishq bolt",
    "fish q vf store",
    "bolt - vfl",
]

# Vehicle / truck / plate detection (these are 0)
VEHICLE_KEYWORDS = [
    "3pl-", "3pl ", "3tpl", "tpl-", "tpl ", "bluepower", "truck", "hiace",
    "hilux", "rider", "keyo fresh logistics",
]
# Named delivery trucks (animals / code names with capacity like 5MT/28MT)
VEHICLE_NAMES = [
    "chui", "duma", "farasi", "kiboko", "kifaru", "mamba", "mwewe", "ndovu",
    "nyati", "simba", "swara", "twiga", "pweza", "skifoc", "paa", "mawenzi",
    "kanga", "njiwa", "mbuni", "kobe", "tai ", "roo truck",
]


def is_vehicle(name):
    low = name.lower()
    for kw in VEHICLE_KEYWORDS:
        if kw in low:
            return True
    for kw in VEHICLE_NAMES:
        if low.startswith(kw) or (" " + kw) in low:
            return True
    # plate-pattern like KDK040Q, KDS 412X, KDP 369K, KCS 559A
    if re.search(r"\bK[A-Z]{2}\s?\d{3}\s?[A-Z]\b", name):
        return True
    if re.search(r"\bKD[A-Z]\d{3}[A-Z]\b", name):
        return True
    if re.search(r"\d+(\.\d+)?\s?mt\b", low):  # capacity like 28MT, 3.5Mt
        return True
    return False


def classify_with_reason(name):
    """Return (zone, reason). zone = 1 (fixed valuation zone) or 0 (non-zone).

    Reasons are tied to Lucia's email logic:
      - Processing is the variance ORIGIN -> 0 (fixed rate applied as stock leaves it).
      - FLC + everything downstream (LCs + branches) inherits the fixed rate -> 1.
      - Non-fish / operational / transit / vehicle / group / test warehouses -> 0.
    """
    low = name.lower()

    if name == "Processing - VFL":
        return 0, "Variance origin: fixed rate applied as fish leaves Processing -> FLC"

    if name in FORCE_ZERO:
        return 0, "Group/parent node or non-stock channel (no fixed-rate fish held here)"

    # Test / dummy
    if "test" in low or "rejected" in low:
        return 0, "Test / rejected / non-operational warehouse"

    # Vehicles & trucks
    if is_vehicle(name):
        return 0, "Vehicle / 3PL truck (transit asset, not a branch stock location)"

    # Internal/operational keywords
    for kw in ZERO_KEYWORDS:
        if kw in low:
            return 0, "Internal/operational store (non-fish inventory: feed, consumables, etc.)"

    # FLC group -> 1
    if low.startswith("flc"):
        return 1, "FLC zone: receives fish from Processing at the fixed rate (email: 'FLC just receives at standard')"

    # Regional LCs and their sub-warehouses -> 1
    if low.startswith("vlc") or low.startswith("mlc") or low.startswith("klc"):
        return 1, "Regional LC downstream of FLC: inherits fixed rate (FLC -> LC transfer, no new variance)"
    if low.startswith("western") or low.startswith("central") or low.startswith("pwani"):
        return 1, "Regional LC/hub downstream of FLC: inherits fixed rate"
    if low.startswith("meru") or low.startswith("mk/t"):
        return 1, "Meru route branch downstream of FLC: inherits fixed rate"
    if low.endswith("region - vfl") or " region - vfl" in low:
        return 1, "Regional distribution node downstream of FLC: inherits fixed rate"

    # Branch route prefixes (Mombasa/Western/Siaya routes) -> 1
    if re.match(r"^(m/t1|w/t3|w/t4|c/t2|c/t3)", low):
        return 1, "Route branch downstream of FLC/LC: inherits fixed rate"

    # Everything else that reaches here is a sales branch / branch spoilage -> 1
    return 1, "Sales branch (or its spoilage/returns): inherits fixed rate, COGS = fixed_rate"


def classify(name):
    """Backward-compatible: return only the zone value."""
    return classify_with_reason(name)[0]


# ------------------------------------------------------------------
# Build the workbook
# ------------------------------------------------------------------

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Classification"

    headers = ["Warehouse Name", "Status", "Fixed Valuation Zone", "Reason / Basis (per email logic)"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    center = Alignment(horizontal="center")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        if col_idx in (2, 3):
            cell.alignment = center

    ones = zeros = 0
    for name, status in WAREHOUSES:
        zone, reason = classify_with_reason(name)
        if zone == 1:
            ones += 1
        else:
            zeros += 1
        ws.append([name, status, zone, reason])
        ws.cell(row=ws.max_row, column=2).alignment = center
        ws.cell(row=ws.max_row, column=3).alignment = center

    # Freeze header and add auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:D" + str(ws.max_row)

    # Auto-size columns (cap the Reason column so it stays readable)
    max_widths = {1: 60, 2: 14, 3: 22, 4: 70}
    for col_idx in range(1, 5):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        width = min(max_len + 3, max_widths.get(col_idx, 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "warehouse_classification.xlsx")
    try:
        wb.save(out_path)
    except PermissionError:
        import time
        alt = os.path.join(out_dir, "warehouse_classification_" + time.strftime("%H%M%S") + ".xlsx")
        wb.save(alt)
        print("NOTE: original file was open/locked; saved to a new file instead.")
        out_path = alt

    print("Saved: " + out_path)
    print("Total warehouses: " + str(len(WAREHOUSES)))
    print("  Fixed Valuation Zone = 1: " + str(ones))
    print("  Fixed Valuation Zone = 0: " + str(zeros))


if __name__ == "__main__":
    main()
